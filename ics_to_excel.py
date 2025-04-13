from icalendar import Calendar
import pandas as pd
from datetime import datetime
import pytz
from calendar import monthrange
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

def convert_ics_to_calendar(ics_file_path, year, month):
    with open(ics_file_path, 'rb') as f:
        cal = Calendar.from_ical(f.read())

    korea_tz = pytz.timezone('Asia/Seoul')

    events = {}
    for component in cal.walk('vevent'):
        start = component.get('dtstart').dt
        end = component.get('dtend').dt if component.get('dtend') else start
        
        if start.tzinfo is None:
            start = pytz.utc.localize(start)
        if end.tzinfo is None:
            end = pytz.utc.localize(end)
        
        start_korea = start.astimezone(korea_tz)
        end_korea = end.astimezone(korea_tz)

        if start_korea.year == year and start_korea.month == month:
            day = start_korea.day
            
            summary = component.get('summary', '')
            if isinstance(summary, bytes):
                summary = summary.decode('utf-8', errors='replace')
            
            event_str = f"{start_korea.strftime('%H:%M')}~{end_korea.strftime('%H:%M')} : {summary}"
            
            if day in events:
                events[day] += '\n' + event_str
            else:
                events[day] = event_str

    first_day = datetime(year, month, 1)
    
    _, days_in_month = monthrange(year, month)
    
    weekdays = ['일', '월', '화', '수', '목', '금', '토']
    
    calendar_dates = [['' for _ in range(7)] for _ in range(6)]
    calendar_events = [['' for _ in range(7)] for _ in range(6)]
    
    first_weekday = first_day.weekday()
    if first_weekday == 6:
        first_weekday = 0
    else:
        first_weekday += 1
    
    day = 1
    for week in range(6):
        for weekday_idx in range(7):
            if week == 0 and weekday_idx < first_weekday:
                calendar_dates[week][weekday_idx] = ''
                calendar_events[week][weekday_idx] = ''
            elif day > days_in_month:
                calendar_dates[week][weekday_idx] = ''
                calendar_events[week][weekday_idx] = ''
            else:
                calendar_dates[week][weekday_idx] = str(day)
                
                if day in events:
                    calendar_events[week][weekday_idx] = events[day]
                else:
                    calendar_events[week][weekday_idx] = ''
                
                day += 1
    
    final_data = []
    
    final_data.append(weekdays)
    
    for week in range(6):
        if any(calendar_dates[week][weekday_idx] != '' for weekday_idx in range(7)):
            final_data.append(calendar_dates[week])
            final_data.append(calendar_events[week])
    
    df = pd.DataFrame(final_data)
    
    excel_file = f'calendar_{year}_{month:02d}.xlsx'
    df.to_excel(excel_file, index=False, header=False)
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    for column in range(1, 8):
        column_letter = get_column_letter(column)
        ws.column_dimensions[column_letter].width = 40
    
    for idx, row in enumerate(ws.iter_rows(), start=1):
        if idx % 2 == 0 or idx == 1:
            ws.row_dimensions[idx].height = 15
        else:
            ws.row_dimensions[idx].height = 113
    
    wb.save(excel_file)
    print(f'달력이 {excel_file}에 저장되었습니다.')

if __name__ == '__main__':
    while True:
        ics_filename = input("gmail 계정 이름을 입력하세요 (예: dev-everyday): ")
        if not ics_filename.endswith('@gmail.com.ics'):
            ics_filename += '@gmail.com.ics'
        
        if os.path.exists(ics_filename):
            break
        else:
            print(f"'{ics_filename}' 파일이 존재하지 않습니다. 다시 입력해주세요.")
    
    year = int(input("년도를 입력하세요 (예: 2025): "))
    month = int(input("월을 입력하세요 (예: 4): "))
    
    convert_ics_to_calendar(ics_filename, year, month) 